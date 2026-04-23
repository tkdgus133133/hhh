#!/usr/bin/env python3
"""datas/ 폴더 전체 → Supabase 이관 스크립트.

실행 전 필수:
  1. Supabase Dashboard → SQL Editor → supabase/schema_sg_tables.sql 실행
  2. .env에 SUPABASE_URL / SUPABASE_KEY 설정 (또는 스크립트 상단 상수 직접 사용)

실행:
  python scripts/migrate_to_supabase.py

이관 대상:
  - products_seed.jsonl      → products (8 KUP 품목)
  - ListingOfRegistered*.csv → products (HSA 등재 5,485건)
  - cancer_incidence.csv     → sg_cancer_incidence
  - API_SP.POP.TOTL*.csv     → sg_world_population
  - SYB67_325*.csv           → sg_health_expenditure
  - 9A706FD_ALL_LATEST.csv   → sg_health_expenditure
  - EML export.xlsx          → sg_who_eml
  - GHED_data.XLSX           → sg_ghed_expenditure (SG 포함 전국)
  - 무역 AX 마스터*.xlsx      → sg_market_targets
  - context_cache.json       → sg_product_context
  - SG sources               → sources
  - PDFs                     → Supabase Storage + sg_documents
"""
from __future__ import annotations

import csv
import json
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env", override=False)
except ImportError:
    pass

SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://oynefikqoibwtfpjlizv.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im95bmVmaWtxb2lid3RmcGpsaXp2Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NjA1NzgwMywiZXhwIjoyMDkxNjMzODAzfQ.eCFcjx7gOhiv7mCyR2RiadndE9d6e6kVOWysHrarZTM")
DATA_DIR = ROOT / "datas"
STORAGE_BUCKET = "sg-documents"

_BATCH = 500  # Supabase 권장 배치 크기


def _patch_openpyxl():
    """openpyxl PageMargins 파서가 빈 문자열을 float으로 변환 시 발생하는 TypeError 억제."""
    try:
        import openpyxl.descriptors.base as _base
        _orig = _base._convert
        def _patched(expected_type, value):
            if value == "" and expected_type is float:
                return 0.0
            return _orig(expected_type, value)
        _base._convert = _patched
    except Exception:
        pass


_patch_openpyxl()


def _client():
    from supabase import create_client
    return create_client(SUPABASE_URL, SUPABASE_KEY)


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _batch_insert(sb, table: str, rows: list[dict], label: str,
                  on_conflict: str | None = None) -> int:
    """배치 단위 upsert. 중복은 무시하고 계속.

    on_conflict: 충돌 키 컬럼명 (콤마 구분). 지정 시 upsert, 미지정 시 insert.
    """
    total = 0
    for i in range(0, len(rows), _BATCH):
        chunk = rows[i: i + _BATCH]
        try:
            if on_conflict:
                sb.table(table).upsert(chunk, on_conflict=on_conflict).execute()
            else:
                sb.table(table).insert(chunk).execute()
            total += len(chunk)
            print(f"  [{label}] {total}/{len(rows)} 적재")
        except Exception as e:
            err_msg = str(e)
            if "23505" in err_msg or "duplicate key" in err_msg:
                # 중복 충돌 → 해당 배치 skip (이미 적재된 데이터)
                print(f"  [{label}] 배치 {i}~{i+len(chunk)} 중복 skip (이미 적재됨)")
            else:
                print(f"  [{label}] 배치 {i}~{i+len(chunk)} 오류: {e}")
        time.sleep(0.05)  # rate limit
    return total


def _safe_storage_name(name: str) -> str:
    """Supabase Storage 허용 문자만 남기도록 파일명 sanitize."""
    import unicodedata
    import re
    # 한글/특수문자 → ASCII로 변환 시도, 실패하면 제거
    name = unicodedata.normalize("NFKD", name)
    name = name.encode("ascii", "ignore").decode("ascii")
    # 허용: 영숫자, 하이픈, 언더스코어, 점, 공백→언더스코어
    name = re.sub(r"[^\w.\-]", "_", name)
    name = re.sub(r"_+", "_", name).strip("_")
    return name


def _check_table_exists(sb, table: str) -> bool:
    try:
        sb.table(table).select("id").limit(1).execute()
        return True
    except Exception:
        try:
            sb.table(table).select("*").limit(1).execute()
            return True
        except Exception:
            return False


# =============================================================================
# 1. SG Sources 등록
# =============================================================================

def migrate_sg_sources(sb) -> None:
    print("\n[1/10] SG Sources 등록")
    sg_sources = [
        {
            "id": "SG:hsa_registry",
            "country": "SG",
            "name": "HSA 등재 의약품 목록",
            "url": "https://www.hsa.gov.sg/therapeutic-products/registered-therapeutic-products",
            "tier": 1,
            "category": "regulator",
            "access_method": "csv",
            "enabled": True,
            "confidence_default": 0.95,
            "rate_limit_qps": None,
            "workflow": "sg_static_migration",
        },
        {
            "id": "SG:kup_pipeline",
            "country": "SG",
            "name": "Korea United Pharm 파이프라인 품목",
            "url": "https://www.kupharma.com",
            "tier": 1,
            "category": "regulator",
            "access_method": "manual",
            "enabled": True,
            "confidence_default": 0.85,
            "rate_limit_qps": None,
            "workflow": "sg_static_migration",
        },
        {
            "id": "SG:who_eml",
            "country": "SG",
            "name": "WHO 필수의약품 목록 (EML 2023)",
            "url": "https://www.who.int/groups/expert-committee-on-selection-and-use-of-essential-medicines/essential-medicines-lists",
            "tier": 1,
            "category": "regulator",
            "access_method": "xlsx",
            "enabled": True,
            "confidence_default": 0.98,
            "rate_limit_qps": None,
            "workflow": "sg_static_migration",
        },
    ]
    try:
        sb.table("sources").upsert(sg_sources, on_conflict="id").execute()
        print(f"  SG 소스 {len(sg_sources)}건 등록 완료")
    except Exception as e:
        print(f"  소스 등록 오류: {e}")


# =============================================================================
# 2. 8 KUP 품목 → products
# =============================================================================

# 분석 메타 (PRODUCT_META 대체)
_KUP_META: dict[str, dict] = {
    "SG_hydrine_hydroxyurea_500": {
        "inn": "hydroxyurea", "atc": "L01XX05",
        "therapeutic_area": "항암(겸상적혈구병, 만성 골수성 백혈병)",
        "hsa_reg": "등재 확인: SIN11083P",
        "key_risk": "세포독성 약물 — 취급·운송 특별 요건. 병원 조달 전용 채널.",
        "product_type": "일반제",
    },
    "SG_gadvoa_gadobutrol_604": {
        "inn": "gadobutrol", "atc": "V08CA09",
        "therapeutic_area": "MRI 조영제 (두개·척추·CE-MRA·간·신장)",
        "hsa_reg": "Gadova 브랜드 미등재 — GADOVIST(레퍼런스) 등재 확인. 브랜드 신규 등록 필요.",
        "key_risk": "브랜드 HSA 신규 등록 필요. macrocyclic GBCA — NSF 위험 최저 등급.",
        "product_type": "일반제",
    },
    "SG_sereterol_activair": {
        "inn": "fluticasone/salmeterol", "atc": "R03AK06",
        "therapeutic_area": "천식·COPD (GINA/GOLD 가이드라인 권고)",
        "hsa_reg": "Sereterol 브랜드 미등재 — SERETIDE(GSK) 등재 확인.",
        "key_risk": "GSK Seretide 특허 만료 여부 확인 필요. 처방전 필요(Rx) 채널.",
        "product_type": "일반제",
    },
    "SG_omethyl_omega3_2g": {
        "inn": "omega-3 acid ethyl esters", "atc": "C10AX06",
        "therapeutic_area": "고중성지방혈증 (Type IV·IIb)",
        "hsa_reg": "미등재 가능성 높음 — omega-3 EE 2g 단독제 HSA CSV 0건. 신규 NDA 필요.",
        "key_risk": "한국 최초 2g 단일캡슐 개량신약. REDUCE-IT 근거 보유.",
        "product_type": "개량신약 (IMD)",
    },
    "SG_rosumeg_combigel": {
        "inn": "rosuvastatin/omega-3 acid ethyl esters", "atc": "C10BA06",
        "therapeutic_area": "이상지질혈증 복합치료 (Type IIb)",
        "hsa_reg": "미등재 확인 — rosuvastatin+omega-3 복합제 HSA CSV 0건.",
        "key_risk": "HOPE-3 근거 (MACE 24% 감소). 복합제 HSA 별도 등재 요건.",
        "product_type": "개량신약 (IMD)",
    },
    "SG_atmeg_combigel": {
        "inn": "atorvastatin/omega-3 acid ethyl esters", "atc": "C10BA05",
        "therapeutic_area": "이상지질혈증 복합치료 (Type IIb)",
        "hsa_reg": "미등재 확인 — atorvastatin+omega-3 복합제 HSA CSV 0건.",
        "key_risk": "ATOM 3상 근거 (non-HDL-C 5% 추가 감소). 복합제 HSA 별도 등재 요건.",
        "product_type": "개량신약 (IMD)",
    },
    "SG_ciloduo_cilosta_rosuva": {
        "inn": "cilostazol/rosuvastatin", "atc": "B01AC23",
        "therapeutic_area": "말초동맥질환·이상지질혈증 복합치료",
        "hsa_reg": "성분 미등재 — cilostazol 단독 HSA CSV 0건. 성분 레벨 신규 NDA Full 요구.",
        "key_risk": "cilostazol 성분 자체 HSA 미등재. 아시아 외 승인 데이터 부족.",
        "product_type": "개량신약 (IMD)",
    },
    "SG_gastiin_cr_mosapride": {
        "inn": "mosapride", "atc": "A03FA05",
        "therapeutic_area": "위장관 운동 촉진 (기능성 소화불량)",
        "hsa_reg": "성분+제품 모두 미등재 — mosapride HSA CSV 0건. NDA Full + 임상 근거 필요.",
        "key_risk": "MARS 3상 비열등성 근거 보유. mosapride 성분 자체 HSA 미등재.",
        "product_type": "개량신약 (IMD)",
    },
}


def migrate_kup_products(sb) -> None:
    print("\n[2/10] 8 KUP 품목 → products")
    seed_path = DATA_DIR / "static" / "products_seed.jsonl"
    if not seed_path.exists():
        print(f"  SKIP: {seed_path} 없음")
        return

    existing = sb.table("products").select("id", count="exact").eq("country", "SG").eq("source_name", "SG:kup_pipeline").execute()
    if getattr(existing, "count", None) and existing.count > 0:
        print(f"  SKIP: 이미 {existing.count}건 존재 (중복 방지)")
        return

    rows = []
    with open(seed_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            seed = json.loads(line)
            pid = seed["product_id"]
            meta = _KUP_META.get(pid, {})
            rows.append({
                "country": "SG",
                "product_id": pid,
                "trade_name": seed.get("trade_name", ""),
                "active_ingredient": seed.get("scientific_name", ""),
                "inn_name": meta.get("inn", seed.get("scientific_name", "")),
                "strength": seed.get("strength"),
                "dosage_form": seed.get("dosage_form"),
                "manufacturer": "Korea United Pharm. Inc.",
                "market_segment": seed.get("market_segment", "retail"),
                "registration_number": seed.get("regulatory_id"),
                "source_name": "SG:kup_pipeline",
                "source_url": f"https://www.kupharma.com/products/{pid}",
                "source_tier": 1,
                "confidence": float(seed.get("confidence", 0.8)),
                "country_specific": {
                    "atc": meta.get("atc"),
                    "therapeutic_area": meta.get("therapeutic_area"),
                    "hsa_reg": meta.get("hsa_reg"),
                    "key_risk": meta.get("key_risk"),
                    "product_type": meta.get("product_type"),
                },
                "raw_payload": seed.get("raw_payload"),
                "crawled_at": _now(),
            })

    _batch_insert(sb, "products", rows, "KUP 품목", on_conflict="country,source_name,source_url")


# =============================================================================
# 3. HSA CSV → products (5,485건)
# =============================================================================

def migrate_hsa_registry(sb) -> None:
    print("\n[3/10] HSA 등재 의약품 → products")
    hsa_path = DATA_DIR / "ListingofRegisteredTherapeuticProducts.csv"
    if not hsa_path.exists():
        print(f"  SKIP: {hsa_path} 없음")
        return

    existing = sb.table("products").select("id", count="exact").eq("country", "SG").eq("source_name", "SG:hsa_registry").execute()
    if getattr(existing, "count", None) and existing.count > 0:
        print(f"  SKIP: 이미 {existing.count}건 존재 (중복 방지)")
        return

    rows = []
    with open(hsa_path, encoding="utf-8", errors="replace") as f:
        for row in csv.DictReader(f):
            lic = (row.get("licence_no") or "").strip()
            if not lic:
                continue
            fc = row.get("forensic_classification", "")
            seg = "retail" if "General" in fc or "Pharmacy" in fc else "retail"

            rows.append({
                "country": "SG",
                "product_id": lic,
                "trade_name": (row.get("product_name") or "").strip(),
                "active_ingredient": (row.get("active_ingredients") or "").replace("&&", " + "),
                "inn_name": (row.get("active_ingredients") or "").split("&&")[0].strip(),
                "strength": (row.get("strength") or "").strip() or None,
                "dosage_form": (row.get("dosage_form") or "").strip().lower() or None,
                "manufacturer": (row.get("manufacturer") or "").strip() or None,
                "market_segment": seg,
                "registration_number": lic,
                "source_name": "SG:hsa_registry",
                # 각 licence_no를 URL에 포함 → dedup unique constraint 통과
                "source_url": f"https://www.hsa.gov.sg/therapeutic-products/registered-therapeutic-products?licence_no={lic}",
                "source_tier": 1,
                "confidence": 0.95,
                "country_specific": {
                    "license_holder": (row.get("license_holder") or "").strip(),
                    "forensic_classification": fc,
                    "atc_code": (row.get("atc_code") or "").strip(),
                    "route_of_administration": (row.get("route_of_administration") or "").strip(),
                    "country_of_manufacturer": (row.get("country_of_manufacturer") or "").strip(),
                    "approval_date": (row.get("approval_d") or "").strip()[:10],
                },
                "raw_payload": None,
                "crawled_at": _now(),
            })

    print(f"  HSA 총 {len(rows)}건 적재 중…")
    _batch_insert(sb, "products", rows, "HSA", on_conflict="country,source_name,source_url")


# =============================================================================
# 4. Cancer incidence → sg_cancer_incidence
# =============================================================================

def migrate_cancer_incidence(sb) -> None:
    print("\n[4/10] 암 발생률 → sg_cancer_incidence")
    if not _check_table_exists(sb, "sg_cancer_incidence"):
        print("  SKIP: sg_cancer_incidence 테이블 없음 — SQL Editor에서 schema 먼저 실행")
        return

    path = DATA_DIR / "dataset-inc-both-sexes-in-2022-singapore.csv"
    if not path.exists():
        print(f"  SKIP: {path} 없음")
        return

    rows = []
    with open(path, encoding="utf-8", errors="replace") as f:
        for r in csv.DictReader(f):
            rows.append({
                "cancer_code": r.get("Cancer code"),
                "icd_code": r.get("ICD Code"),
                "label": r.get("Label", ""),
                "sex": int(r["Sex"]) if r.get("Sex", "").isdigit() else None,
                "number": _int(r.get("Number")),
                "ui_low": _int(r.get("95% UI low")),
                "ui_high": _int(r.get("95% UI high")),
                "asr_world": _float(r.get("ASR (World)")),
                "crude_rate": _float(r.get("Crude rate")),
                "cumulative_risk": _float(r.get("Cumulative risk")),
                "data_year": 2022,
            })
    _batch_insert(sb, "sg_cancer_incidence", rows, "암 발생률")


# =============================================================================
# 5. World population → sg_world_population
# =============================================================================

def migrate_world_population(sb) -> None:
    print("\n[5/10] 세계 인구 → sg_world_population")
    if not _check_table_exists(sb, "sg_world_population"):
        print("  SKIP: sg_world_population 테이블 없음 — SQL Editor에서 schema 먼저 실행")
        return

    existing = sb.table("sg_world_population").select("id", count="exact").execute()
    if getattr(existing, "count", None) and existing.count > 0:
        print(f"  SKIP: 이미 {existing.count}건 존재 (중복 방지)")
        return

    path = DATA_DIR / "API_SP.POP.TOTL_DS2_en_csv_v2_58.csv"
    if not path.exists():
        print(f"  SKIP: {path} 없음")
        return

    rows = []
    with open(path, encoding="utf-8", errors="replace") as f:
        content = f.read()

    # 헤더 행 찾기 (Country Name, Country Code, Indicator Name, Indicator Code, 1960, 1961, ...)
    for line in content.split("\n"):
        if "Country Name" in line:
            headers = next(csv.reader([line]))
            year_cols = {i: int(h) for i, h in enumerate(headers) if h.isdigit() and 1960 <= int(h) <= 2030}
            break
    else:
        print("  SKIP: 헤더를 찾을 수 없음")
        return

    for line in content.split("\n"):
        if not line.strip() or "Country Name" in line:
            continue
        parts = next(csv.reader([line]))
        if len(parts) < 4:
            continue
        cname, ccode = parts[0].strip('"'), parts[1].strip('"')
        if not ccode or len(ccode) != 3:
            continue
        for col_idx, year in year_cols.items():
            if col_idx >= len(parts):
                continue
            val = parts[col_idx].strip().strip('"')
            if val:
                try:
                    rows.append({
                        "country_name": cname,
                        "country_code": ccode,
                        "year": year,
                        "population": int(float(val)),
                    })
                except ValueError:
                    pass

    print(f"  인구 데이터 총 {len(rows)}건 적재 중…")
    _batch_insert(sb, "sg_world_population", rows, "세계 인구", on_conflict="country_code,year")


# =============================================================================
# 6. SYB67 보건 지출 → sg_health_expenditure
# =============================================================================

def migrate_health_expenditure(sb) -> None:
    print("\n[6/10] 보건 지출 → sg_health_expenditure")
    if not _check_table_exists(sb, "sg_health_expenditure"):
        print("  SKIP: sg_health_expenditure 테이블 없음 — SQL Editor에서 schema 먼저 실행")
        return

    path = DATA_DIR / "SYB67_325_202411_Expenditure on health.csv"
    if not path.exists():
        print(f"  SKIP: {path} 없음")
        return

    rows = []
    with open(path, encoding="utf-8", errors="replace") as f:
        # 첫 줄은 파일 타이틀(T11,Expenditure on health,...)이므로 건너뜀
        first = f.readline()
        # 첫 줄이 실제 헤더(Region/Country/Area 포함)면 되돌려 읽기
        if "Region/Country/Area" in first or "Country" in first:
            f.seek(0)
        reader = csv.DictReader(f)
        for r in reader:
            country = (r.get("Region/Country/Area") or r.get("T08") or "").strip()
            year_raw = (r.get("Year") or r.get("T09") or "").strip()
            series = (r.get("Series") or r.get("T10") or "").strip()
            value_raw = (r.get("Value") or r.get("T11") or "").strip()
            if not country or not year_raw.isdigit():
                continue
            rows.append({
                "country_or_area": country,
                "year": int(year_raw),
                "series": series,
                "value": _float(value_raw),
                "footnotes": r.get("Footnotes", ""),
                "source": r.get("Source", ""),
            })

    print(f"  보건 지출 총 {len(rows)}건 적재 중…")
    _batch_insert(sb, "sg_health_expenditure", rows, "보건 지출")


# =============================================================================
# 7. GHED XLSX → sg_ghed_expenditure
# =============================================================================

def migrate_ghed(sb) -> None:
    print("\n[7/10] GHED 보건 지출 → sg_ghed_expenditure")
    if not _check_table_exists(sb, "sg_ghed_expenditure"):
        print("  SKIP: sg_ghed_expenditure 테이블 없음 — SQL Editor에서 schema 먼저 실행")
        return

    # 이미 적재된 데이터가 있으면 upsert로 처리 (중복 skip)
    try:
        existing = sb.table("sg_ghed_expenditure").select("id", count="exact").limit(1).execute()
        existing_count = existing.count or 0
        if existing_count > 0:
            print(f"  이미 {existing_count:,}건 적재됨 → upsert 모드로 실행 (중복 skip)")
    except Exception:
        existing_count = 0

    path = DATA_DIR / "GHED_data.XLSX"
    if not path.exists():
        print(f"  SKIP: {path} 없음")
        return

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), data_only=True)
        ws = wb["Data"] if "Data" in wb.sheetnames else wb.active
    except Exception as e:
        print(f"  GHED 파일 열기 오류: {e}")
        return

    headers = None
    rows = []
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(c).strip() if c else "" for c in row]
            continue
        if not any(row):
            continue
        r = dict(zip(headers, row))
        # GHED 컬럼명: location(국가명), code(ISO3), year
        country = str(r.get("location") or r.get("country") or "").strip()
        code = str(r.get("code") or r.get("country_code") or "").strip()
        year_val = r.get("year") or r.get("Year")
        if not country or not year_val:
            continue
        try:
            year = int(year_val)
        except (ValueError, TypeError):
            continue

        # GHED의 모든 지표 컬럼을 개별 행으로 변환
        skip = {"location", "country", "code", "year", "region", "income", "country_code", "country_name"}
        for col, val in r.items():
            if col in skip or not col or val is None or val == "":
                continue
            try:
                fval = float(val)
            except (ValueError, TypeError):
                continue
            rows.append({
                "country": country,
                "country_code": code or None,
                "year": year,
                "indicator_code": col,
                "indicator_name": col,
                "value": fval,
            })
        if len(rows) >= 50000:  # 메모리 절약을 위해 중간 플러시
            _batch_insert(sb, "sg_ghed_expenditure", rows, "GHED",
                          on_conflict="country_code,year,indicator_code")
            rows = []

    if rows:
        _batch_insert(sb, "sg_ghed_expenditure", rows, "GHED",
                      on_conflict="country_code,year,indicator_code")

    wb.close()


# =============================================================================
# 8. WHO EML → sg_who_eml
# =============================================================================

def migrate_who_eml(sb) -> None:
    print("\n[8/10] WHO EML → sg_who_eml")
    if not _check_table_exists(sb, "sg_who_eml"):
        print("  SKIP: sg_who_eml 테이블 없음 — SQL Editor에서 schema 먼저 실행")
        return

    path = DATA_DIR / "EML export.xlsx"
    if not path.exists():
        print(f"  SKIP: {path} 없음")
        return

    try:
        import pandas as pd
        df = pd.read_excel(str(path), engine="openpyxl")
    except Exception as e:
        print(f"  EML 파일 열기 오류: {e}")
        return

    rows = []
    def _s(v):
        s = str(v).strip()
        return None if s in ("", "nan", "None") else s

    for _, r in df.iterrows():
        # 실제 컬럼명: 'Medicine name', 'ATC codes', 'EML section', 'Formulations', 'Indication'
        inn = str(
            r.get("Medicine name") or r.get("INN") or r.get("name") or r.get("Medicine") or ""
        ).strip()
        if not inn or inn == "nan":
            continue
        rows.append({
            "inn_name": inn,
            "atc_code": _s(r.get("ATC codes") or r.get("ATC code") or r.get("atc_code")),
            "dosage_form": _s(r.get("Formulations") or r.get("Dosage form") or r.get("dosage_form")),
            "strength": _s(r.get("Strength") or r.get("strength")),
            "section_code": _s(r.get("Section") or r.get("section_code")),
            "section_name": _s(r.get("EML section") or r.get("Section name") or r.get("section_name")),
            "eml_type": _s(r.get("Status") or r.get("Type") or r.get("type")),
            "indication": _s(r.get("Indication") or r.get("indication")),
            "notes": _s(r.get("Combined with") or r.get("Notes") or r.get("notes")),
            "eml_year": 2023,
            "raw_payload": {k: str(v) for k, v in r.items() if str(v) not in ("", "nan", "None")},
        })

    print(f"  WHO EML 총 {len(rows)}건 적재 중…")
    _batch_insert(sb, "sg_who_eml", rows, "WHO EML")


# =============================================================================
# 9. context_cache.json → sg_product_context
# =============================================================================

def migrate_product_context(sb) -> None:
    print("\n[9/10] context_cache.json → sg_product_context")
    if not _check_table_exists(sb, "sg_product_context"):
        print("  SKIP: sg_product_context 테이블 없음 — SQL Editor에서 schema 먼저 실행")
        return

    existing = sb.table("sg_product_context").select("id", count="exact").execute()
    if getattr(existing, "count", None) and existing.count > 0:
        print(f"  SKIP: 이미 {existing.count}건 존재 (중복 방지)")
        return

    path = DATA_DIR / "static" / "context_cache.json"
    if not path.exists():
        print(f"  SKIP: {path} 없음")
        return

    cache = json.loads(path.read_text(encoding="utf-8"))
    rows = []
    for pid, ctx in cache.items():
        rows.append({
            "product_id": pid,
            "hsa_matches": ctx.get("hsa_matches", []),
            "hsa_registered": ctx.get("hsa_registered", False),
            "competitor_count": ctx.get("competitor_count", 0),
            "prescription_only": ctx.get("prescription_only", True),
            "pdf_snippets": ctx.get("pdf_snippets", []),
            "brochure_snippets": ctx.get("brochure_snippets", []),
            "regulatory_summary": ctx.get("regulatory_summary", ""),
            "built_at": ctx.get("built_at") or _now(),
            "updated_at": _now(),
        })
    _batch_insert(sb, "sg_product_context", rows, "product_context", on_conflict="product_id")


# =============================================================================
# 10. PDFs → Supabase Storage + sg_documents
# =============================================================================

def migrate_pdfs(sb) -> None:
    print("\n[10/10] PDF → Supabase Storage")
    if not _check_table_exists(sb, "sg_documents"):
        print("  SKIP: sg_documents 테이블 없음 — SQL Editor에서 schema 먼저 실행")
        return

    # 이미 적재된 파일 목록 조회 → storage_path 기준으로 중복 skip
    try:
        existing_res = sb.table("sg_documents").select("storage_path").execute()
        already_uploaded = {r["storage_path"] for r in (existing_res.data or [])}
        if already_uploaded:
            print(f"  이미 {len(already_uploaded)}건 존재 → 나머지만 추가 업로드")
    except Exception:
        already_uploaded = set()

    # 버킷 생성 (이미 있으면 skip)
    try:
        sb.storage.create_bucket(STORAGE_BUCKET, options={"public": False})
        print(f"  Storage 버킷 '{STORAGE_BUCKET}' 생성")
    except Exception:
        print(f"  Storage 버킷 '{STORAGE_BUCKET}' 이미 존재")

    _PDF_META: list[dict] = [
        # 규제/가이드라인
        {"path": DATA_DIR / "singapore_regulation.pdf",                        "category": "regulation",  "label": "HSA 제품 등록 가이드", "pid": None},
        {"path": DATA_DIR / "guidance-on-therapeutic-product-registration-in-singapore.pdf", "category": "regulation", "label": "싱가포르 의약품 등록 가이드라인", "pid": None},
        {"path": DATA_DIR / "Review-Pricing-policies.pdf",                     "category": "regulation",  "label": "가격 정책 리뷰", "pid": None},
        {"path": DATA_DIR / "싱가포르_1공정_크롤링_보고서_v3_헌법준수.pdf",        "category": "report",      "label": "싱가포르 1공정 크롤링 보고서 v3", "pid": None},
        {"path": DATA_DIR / "252026싱가포르진출전략.pdf",                         "category": "strategy",    "label": "KOTRA 싱가포르 진출 전략 2026", "pid": None},
        {"path": DATA_DIR / "zjma-7-1601060.pdf",                              "category": "paper",       "label": "의약품 시장 논문", "pid": None},
        {"path": DATA_DIR / "WHO-MHP-HPS-EML-2023.02-eng.pdf",                "category": "regulation",  "label": "WHO EML 2023", "pid": None},
        # 브로슈어
        {"path": DATA_DIR / "basic_files" / "Brochure_Hydrine 500mg cap.pdf", "category": "brochure", "label": "Hydrine 브로슈어", "pid": "SG_hydrine_hydroxyurea_500"},
        {"path": DATA_DIR / "basic_files" / "[KUPZINE]Gadova 604.72mg inj..pdf", "category": "brochure", "label": "Gadova 브로슈어", "pid": "SG_gadvoa_gadobutrol_604"},
        {"path": DATA_DIR / "basic_files" / "제품교육_Sereterol Activair.pdf", "category": "brochure", "label": "Sereterol Activair 브로슈어", "pid": "SG_sereterol_activair"},
        {"path": DATA_DIR / "basic_files" / "Brochure_Omethyl Cutielet.pdf",  "category": "brochure", "label": "Omethyl 브로슈어", "pid": "SG_omethyl_omega3_2g"},
        {"path": DATA_DIR / "basic_files" / "Brochure_Rosumeg Combigel.pdf",  "category": "brochure", "label": "Rosumeg 브로슈어", "pid": "SG_rosumeg_combigel"},
        {"path": DATA_DIR / "basic_files" / "Brochure_Atmeg Combigel.pdf",    "category": "brochure", "label": "Atmeg 브로슈어", "pid": "SG_atmeg_combigel"},
        {"path": DATA_DIR / "basic_files" / "[KUPZINE]Ciloduo 100mg 200mg tab..pdf", "category": "brochure", "label": "Ciloduo 브로슈어", "pid": "SG_ciloduo_cilosta_rosuva"},
        {"path": DATA_DIR / "basic_files" / "Brochure_Gastiin CR.pdf",        "category": "brochure", "label": "Gastiin CR 브로슈어", "pid": "SG_gastiin_cr_mosapride"},
        # 논문
        {"path": DATA_DIR / "papers" / "2104.03154v3.pdf",    "category": "paper", "label": "논문 2104.03154", "pid": None},
        {"path": DATA_DIR / "papers" / "2401.13919v4.pdf",    "category": "paper", "label": "논문 2401.13919", "pid": None},
        {"path": DATA_DIR / "papers" / "2403.14151v2.pdf",    "category": "paper", "label": "논문 2403.14151", "pid": None},
        {"path": DATA_DIR / "papers" / "2411.15100v3.pdf",    "category": "paper", "label": "논문 2411.15100", "pid": None},
        {"path": DATA_DIR / "papers" / "2508.16571v3.pdf",    "category": "paper", "label": "논문 2508.16571", "pid": None},
        {"path": DATA_DIR / "papers" / "2604.02276v1.pdf",    "category": "paper", "label": "논문 2604.02276", "pid": None},
        {"path": DATA_DIR / "papers" / "ESCoE-TR-12.pdf",     "category": "paper", "label": "ESCoE TR-12", "pid": None},
        {"path": DATA_DIR / "papers" / "NIESR-DP-523-4.pdf",  "category": "paper", "label": "NIESR DP-523-4", "pid": None},
        {"path": DATA_DIR / "papers" / "main.pdf",            "category": "paper", "label": "논문 main", "pid": None},
        # AX 마스터 캡스톤
        {"path": DATA_DIR / "basic_files" / "1. 무역 AX 마스터 캡스톤 프로젝트_시장조사 희망 대상.xlsx", "category": "market", "label": "AX 마스터 캡스톤 시장조사", "pid": None},
    ]

    doc_rows = []
    for item in _PDF_META:
        fpath = item["path"]
        if not fpath.exists():
            print(f"  SKIP (파일 없음): {fpath.name}")
            continue

        safe_name = _safe_storage_name(fpath.stem) + fpath.suffix.lower()
        storage_path = f"{item['category']}/{safe_name}"

        # 이미 DB에 등록된 파일은 skip
        if storage_path in already_uploaded:
            print(f"  SKIP (이미 적재): {safe_name}")
            continue

        try:
            with open(fpath, "rb") as f:
                content = f.read()
            ct = "application/pdf" if fpath.suffix.lower() == ".pdf" else "application/octet-stream"
            sb.storage.from_(STORAGE_BUCKET).upload(
                path=storage_path,
                file=content,
                file_options={"content-type": ct, "upsert": "true"},
            )
            doc_rows.append({
                "filename": safe_name,
                "storage_path": storage_path,
                "bucket": STORAGE_BUCKET,
                "category": item["category"],
                "product_id": item["pid"],
                "label": item["label"],
                "file_size_bytes": len(content),
            })
            print(f"  업로드: {safe_name} ({len(content)//1024}KB)")
        except Exception as e:
            print(f"  업로드 실패 {fpath.name}: {str(e)[:80]}")

    if doc_rows:
        # already_uploaded 로 중복 제거 완료 → 일반 insert (constraint 불필요)
        _batch_insert(sb, "sg_documents", doc_rows, "sg_documents")


# =============================================================================
# 진행 현황 리포트
# =============================================================================

def _qcount(sb, table: str, col: str | None = None, val: str | None = None) -> int:
    """테이블의 현재 적재 건수 조회. 테이블 없으면 -1 반환."""
    try:
        q = sb.table(table).select("id", count="exact")
        if col and val:
            q = q.eq(col, val)
        res = q.limit(1).execute()
        return res.count if res.count is not None else 0
    except Exception:
        try:
            q2 = sb.table(table).select("*", count="exact")
            if col and val:
                q2 = q2.eq(col, val)
            res2 = q2.limit(1).execute()
            return res2.count if res2.count is not None else 0
        except Exception:
            return -1


def _csv_count(path: Path) -> int:
    """CSV 파일의 데이터 행 수(헤더 제외)를 반환."""
    try:
        with open(path, encoding="utf-8", errors="replace") as f:
            return sum(1 for _ in csv.DictReader(f))
    except Exception:
        return 0


def _pop_db_row_count(path: Path) -> int:
    """인구 CSV → DB 행 수 추산 (국가 × 연도별 non-empty 값)."""
    try:
        with open(path, encoding="utf-8", errors="replace") as f:
            content = f.read()
        headers = None
        year_cols: dict[int, int] = {}
        for line in content.split("\n"):
            if "Country Name" in line:
                hdrs = next(csv.reader([line]))
                year_cols = {i: int(h) for i, h in enumerate(hdrs)
                             if h.isdigit() and 1960 <= int(h) <= 2030}
                headers = hdrs
                break
        if not headers:
            return 0
        count = 0
        for line in content.split("\n"):
            if not line.strip() or "Country Name" in line:
                continue
            parts = next(csv.reader([line]))
            if len(parts) < 4:
                continue
            ccode = parts[1].strip('"')
            if not ccode or len(ccode) != 3:
                continue
            for col_idx in year_cols:
                if col_idx < len(parts) and parts[col_idx].strip().strip('"'):
                    count += 1
        return count
    except Exception:
        return 0


def show_migration_status(sb) -> None:
    """현재 이관 진행 현황 요약 출력."""
    W = 70
    print("\n" + "━" * W)
    print("  📊  Supabase 이관 진행 현황")
    print("━" * W)
    print(f"  {'단계':<4} {'대상':<26} {'적재됨':>10} {'원본전체':>10} {'남은건수':>10}  상태")
    print("  " + "─" * (W - 2))

    def _icon(done: int, total: int | None) -> str:
        if done == -1:
            return "⚠️  테이블없음"
        if total is None:
            return "⏳ 진행 중" if done > 0 else "❌ 미시작"
        if total == 0:
            return "⚠️  원본없음"
        if done >= total:
            return "✅ 완료"
        if done > 0:
            return f"⏳ 진행 중 ({done/total*100:.1f}%)"
        return "❌ 미시작"

    def _row(step: str, label: str, done: int, total: int | None) -> None:
        done_s = f"{done:,}" if done >= 0 else "조회실패"
        if total is None:
            total_s, remain_s = "동적계산", "-"
        elif total == 0:
            total_s, remain_s = "파일없음", "-"
        else:
            total_s = f"{total:,}"
            remain_s = f"{max(0, total - done):,}" if done >= 0 else "-"
        icon = _icon(done, total)
        print(f"  {step:<4} {label:<26} {done_s:>10} {total_s:>10} {remain_s:>10}  {icon}")

    # ── 원본 파일 건수 계산 (CSV는 빠름) ──────────────────────────────────
    hsa_total  = _csv_count(DATA_DIR / "ListingofRegisteredTherapeuticProducts.csv")
    canc_total = _csv_count(DATA_DIR / "dataset-inc-both-sexes-in-2022-singapore.csv")
    he_total   = _csv_count(DATA_DIR / "SYB67_325_202411_Expenditure on health.csv")
    pop_total  = _pop_db_row_count(DATA_DIR / "API_SP.POP.TOTL_DS2_en_csv_v2_58.csv")

    # PDF 목록: sg_documents 에 등록될 파일 수 (파일 존재하는 것만)
    _PDF_PATHS = [
        DATA_DIR / "singapore_regulation.pdf",
        DATA_DIR / "guidance-on-therapeutic-product-registration-in-singapore.pdf",
        DATA_DIR / "Review-Pricing-policies.pdf",
        DATA_DIR / "싱가포르_1공정_크롤링_보고서_v3_헌법준수.pdf",
        DATA_DIR / "252026싱가포르진출전략.pdf",
        DATA_DIR / "zjma-7-1601060.pdf",
        DATA_DIR / "WHO-MHP-HPS-EML-2023.02-eng.pdf",
        DATA_DIR / "basic_files" / "Brochure_Hydrine 500mg cap.pdf",
        DATA_DIR / "basic_files" / "[KUPZINE]Gadova 604.72mg inj..pdf",
        DATA_DIR / "basic_files" / "제품교육_Sereterol Activair.pdf",
        DATA_DIR / "basic_files" / "Brochure_Omethyl Cutielet.pdf",
        DATA_DIR / "basic_files" / "Brochure_Rosumeg Combigel.pdf",
        DATA_DIR / "basic_files" / "Brochure_Atmeg Combigel.pdf",
        DATA_DIR / "basic_files" / "[KUPZINE]Ciloduo 100mg 200mg tab..pdf",
        DATA_DIR / "basic_files" / "Brochure_Gastiin CR.pdf",
        DATA_DIR / "papers" / "2104.03154v3.pdf",
        DATA_DIR / "papers" / "2401.13919v4.pdf",
        DATA_DIR / "papers" / "2403.14151v2.pdf",
        DATA_DIR / "papers" / "2411.15100v3.pdf",
        DATA_DIR / "papers" / "2508.16571v3.pdf",
        DATA_DIR / "papers" / "2604.02276v1.pdf",
        DATA_DIR / "papers" / "ESCoE-TR-12.pdf",
        DATA_DIR / "papers" / "NIESR-DP-523-4.pdf",
        DATA_DIR / "papers" / "main.pdf",
        DATA_DIR / "basic_files" / "1. 무역 AX 마스터 캡스톤 프로젝트_시장조사 희망 대상.xlsx",
    ]
    pdf_total = sum(1 for p in _PDF_PATHS if p.exists())

    # ── Supabase 현재 적재 건수 조회 ──────────────────────────────────────
    _row("1",  "sources (SG)",
         _qcount(sb, "sources"), 3)
    _row("2",  "products (KUP 8품목)",
         _qcount(sb, "products", "source_name", "SG:kup_pipeline"), 8)
    _row("3",  "products (HSA)",
         _qcount(sb, "products", "source_name", "SG:hsa_registry"), hsa_total)
    _row("4",  "sg_cancer_incidence",
         _qcount(sb, "sg_cancer_incidence"), canc_total)
    _row("5",  "sg_world_population",
         _qcount(sb, "sg_world_population"), pop_total if pop_total else None)
    _row("6",  "sg_health_expenditure",
         _qcount(sb, "sg_health_expenditure"), he_total)
    ghed_done = _qcount(sb, "sg_ghed_expenditure")
    _row("7",  "sg_ghed_expenditure",
         ghed_done, None)   # 전체 건수는 null 제외 동적 계산이라 실행 후에 확정
    _row("8",  "sg_who_eml",
         _qcount(sb, "sg_who_eml"), None)   # XLSX → 실행 전 건수 불확실
    _row("9",  "sg_product_context",
         _qcount(sb, "sg_product_context"), 8)
    _row("10", "sg_documents (PDF)",
         _qcount(sb, "sg_documents"), pdf_total)

    print("━" * W)
    if ghed_done > 0:
        print(f"  ※ GHED: Excel 4,612행 × 최대 4,115개 지표 → null 제외 실제 건수는 적재 완료 후 확정됩니다.")
    print()


# =============================================================================
# 헬퍼
# =============================================================================

def _int(v: Any) -> int | None:
    try:
        return int(float(str(v).replace(",", "")))
    except (ValueError, TypeError):
        return None


def _float(v: Any) -> float | None:
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return None


# =============================================================================
# Main
# =============================================================================

def main() -> None:
    print("=" * 60)
    print("Supabase 이관 시작")
    print(f"URL: {SUPABASE_URL}")
    print("=" * 60)

    sb = _client()

    # ── 진행 현황 먼저 출력 ────────────────────────────────────────────────
    show_migration_status(sb)

    # 필수 SG 테이블 존재 여부 확인
    required = ["sg_cancer_incidence", "sg_world_population", "sg_health_expenditure",
                "sg_ghed_expenditure", "sg_who_eml", "sg_documents",
                "sg_product_context", "sg_market_targets"]
    missing = [t for t in required if not _check_table_exists(sb, t)]
    if missing:
        print(f"\n⚠️  다음 테이블이 없습니다: {missing}")
        print("   → Supabase Dashboard > SQL Editor에서 supabase/schema_sg_tables.sql을 먼저 실행하세요.")
        print("   → 해당 테이블 이관은 SKIP되고 나머지는 계속 진행됩니다.\n")

    migrate_sg_sources(sb)
    migrate_kup_products(sb)
    migrate_hsa_registry(sb)
    migrate_cancer_incidence(sb)
    migrate_world_population(sb)
    migrate_health_expenditure(sb)
    migrate_ghed(sb)
    migrate_who_eml(sb)
    migrate_product_context(sb)
    migrate_pdfs(sb)

    print("\n" + "=" * 60)
    print("이관 완료!")
    print("=" * 60)


if __name__ == "__main__":
    main()
